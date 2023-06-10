from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import time
import openpyxl
import os

def scrape_posts(link, topic, sheet):
    # Configure Selenium WebDriver (you need to download the appropriate driver for your browser)
    geckodriver_path = './webdriver/geckodriver.exe'  # Update with the correct file name and extension
    browser = webdriver.Firefox(executable_path=geckodriver_path)

    # Open the webpage
    browser.get(link)

    div_elements = [] # Emply list to take all articles which would be scraped
    # Perform scrolling
    first_run = 1
    while True:
        # Wait for some time for the new content to load
        time.sleep(2)  # Wait for 2 seconds

        html_source = browser.page_source
        soup = BeautifulSoup(html_source, 'html.parser')

        article_list = soup.find('div', class_='post-stream').find_all('div',
                                                                       class_=lambda value: value and value.startswith('topic-post clearfix'))
        for article in article_list:
            if article not in div_elements:
                div_elements.append(article)

        # Execute JavaScript to scroll to the bottom
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # Check if the bottom of the page is reached
        current_height = browser.execute_script("return document.body.scrollHeight")
        if first_run == 1:
            prev_height = 0
            first_run = 0
        if current_height == prev_height:
            # print(f"Since current length {current_height} and previous length {prev_height} are the same, exiting from loop")
            break  # Reached the end of the page, exit the loop
        else:
            # print(f"Since the current length {current_height}\t is more than prev length - {prev_height}, trying one more scroll")
            prev_height = current_height
            
    # print("Done scrolling and ", "\n", "*" * 15)

    print(f"Number of posts that would be extracted - {len(div_elements)}. Please do check for duplicate posts")

    row = sheet.max_row + 1
    for index, post in enumerate(div_elements):
        post_number = f"Post Number:{index + 1}"
        user_name = post.find('div', class_='names trigger-user-card').text
        content = post.find('div', class_='cooked').text
        post_date = post.find('span', class_='relative-date')['title']
        like_text = post.find('div', class_='double-button').text
        likes = 0 if like_text == "" else int(like_text)
        reply_class = post.find('button', class_='widget-button btn-flat show-replies btn-icon-text')
        reply = '0 Replies' if reply_class is None else reply_class.find('span', class_='d-button-label').text
        
        sheet.cell(row=row, column=1).value = topic
        sheet.cell(row=row, column=2).value = post_number
        sheet.cell(row=row, column=3).value = user_name
        sheet.cell(row=row, column=4).value = content
        sheet.cell(row=row, column=5).value = reply
        sheet.cell(row=row, column=6).value = post_date
        sheet.cell(row=row, column=7).value = likes
        
        row += 1
    print(f"Topic {topic} completed. Moving to next")

    browser.quit()
    
    
directory_path = 'data'
outfile_file = os.path.join(directory_path,'data.xlsx')

# Load the data file
data_workbook = openpyxl.load_workbook(outfile_file)
data_sheet = data_workbook.active

# Create a new workbook for posts
# posts_workbook = openpyxl.Workbook()
posts_sheet = data_workbook.create_sheet(title='Posts')
# create titles

posts_sheet.cell(row=1, column=1).value = 'Title'
posts_sheet.cell(row=1, column=2).value = 'Post number'
posts_sheet.cell(row=1, column=3).value = 'User name'
posts_sheet.cell(row=1, column=4).value = 'Content'
posts_sheet.cell(row=1, column=5).value = 'Replies'
posts_sheet.cell(row=1, column=6).value = 'Posted date'
posts_sheet.cell(row=1, column=7).value = 'Number of likes'

# Iterate through the links in the data file
for row in data_sheet.iter_rows(min_row=2, max_col=7, values_only=True):
    topic, link = row[0], row[6]
    print(f'Topic - {topic} started')
    scrape_posts(link, topic, posts_sheet)
    print(f'Topic - {topic} finished')
    print("*"*50)
    data_workbook.save(outfile_file)
    




