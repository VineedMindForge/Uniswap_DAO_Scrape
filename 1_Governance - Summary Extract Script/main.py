from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl,re,time,os

# Initiate browser driver and open page of interest
geckodriver_path = 'C:/Users/hp/Documents/DAO/webdriver/geckodriver.exe'  # Update with the correct file name and extension
directory_path = 'data'
outfile_file = os.path.join(directory_path,'data.xlsx')

# Check/create output directory
if not os.path.exists(directory_path):
    os.makedirs(directory_path)

browser = webdriver.Firefox(executable_path=geckodriver_path)
browser.get("https://gov.uniswap.org/")

# Infinite scroll till the end of the page
first_run = 1
while True:
    # Execute JavaScript to scroll to the bottom
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    # Wait for some time for the new content to load
    # You can adjust the sleep duration based on the website and your internet speed
    time.sleep(2)  # Wait for 2 seconds

    # Check if the bottom of the page is reached
    current_height = browser.execute_script("return document.body.scrollHeight")
    if first_run ==1:
        prev_height = 0
        first_run =0
    if current_height == prev_height:
        break  # Reached the end of the page, exit the loop
    else:
        print(f"Current Height - {current_height}\t Prev Height - {prev_height}")
        prev_height = current_height
        print('Moving down')

print("Done scrolling", "\n","*"*15)

# Get the HTML source of the page and parse using bs4
html_source = browser.page_source

soup = BeautifulSoup(html_source, 'html.parser')

# Open Workbook object and create titles
sheet = openpyxl.Workbook().active
sheet.cell(row=1, column=1).value = 'Title'
sheet.cell(row=1, column=2).value = 'Description'
sheet.cell(row=1, column=3).value = 'Replies'
sheet.cell(row=1, column=4).value = 'Views'
sheet.cell(row=1, column=5).value = 'First Post'
sheet.cell(row=1, column=6).value = 'Posted'
sheet.cell(row=1, column=7).value = 'Link'

# Perform data extraction using BeautifulSoup methods
data = soup.find('tbody', class_='topic-list-body').find_all('tr')
row = 2
for discussion in data:
    title = discussion.find('td', class_='main-link clearfix topic-list-data').find('span','link-top-line').text.strip()
    try:
        description = discussion.find('a', class_ = 'topic-excerpt').text.strip()
        
    except:
        try:
            description = discussion.find('td', class_='main-link clearfix topic-list-data').find('span',class_='category-name').text.strip()
            
        except:
            description = '*****No valid description extracted*****'
    reply = discussion.find('td').find_next_sibling('td').find_next_sibling('td')['title']
    views_text = discussion.find('td').find_next_sibling('td').find_next_sibling('td').find_next_sibling('td').find('span')['title']
    views = re.findall(r'\d+', views_text)[0]
    dates_test = discussion.find('td').find_next_sibling('td').find_next_sibling('td').find_next_sibling('td').find_next_sibling('td')['title'].strip()
    first_post_date, posted_date = [line.split(': ')[1] for line in dates_test.split('\n')]
    link = "https://gov.uniswap.org"+discussion.find('a', class_='title raw-link raw-topic-link')['href']
      
    # Save data points in specific cells
    sheet.cell(row=row, column=1).value = title
    sheet.cell(row=row, column=2).value = description
    sheet.cell(row=row, column=3).value = reply
    sheet.cell(row=row, column=4).value = views
    sheet.cell(row=row, column=5).value = first_post_date
    sheet.cell(row=row, column=6).value = posted_date
    sheet.cell(row=row, column=7).value = link
    row+=1
    
    sheet.parent.save(outfile_file)

browser.quit()