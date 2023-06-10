from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import time
import openpyxl
import os

def scrape_posts(number,link, topic, sheet):
    
    # Configure Selenium WebDriver (you need to download the appropriate driver for your browser)
    geckodriver_path = './webdriver/geckodriver.exe'  # Update with the correct file name and extension
    browser = webdriver.Firefox(executable_path=geckodriver_path)

    # Open the webpage
    browser.get(link)
    wait = WebDriverWait(browser, 10)
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.sc-1vobn0f-10')))
    time.sleep(10)
       
    html_source = browser.page_source
    soup = BeautifulSoup(html_source,'html.parser') 
    
    
    voting_dates = soup.find_all('div', class_='sc-sx9n2y-0 bqwbXT css-8mokm4')[1].text
    for_votes_element =  soup.find('div', class_= 'sc-bczRLJ sc-nrd8cx-0 sc-nrd8cx-1 sc-1vobn0f-8 hJYFVB fhPvJh frnZMK eAwuXe')
    for_votes = for_votes_element.text.replace('For',"").replace(',',"").replace(' ','') 
    against_votes_element = soup.find('div', class_='sc-1kykgp9-2 sc-5nm9ef-0 sc-1vobn0f-4 jdTKGL gqCodE kWytlN').find_next_sibling()
    against_votes = against_votes_element.text.replace('Against',"").replace(',',"").replace(' ','') 
    details_text = soup.find('div',class_='sc-1vobn0f-9 eaveHp').text
    description_text = soup.find('div',class_='sc-1vobn0f-7 gPKgVe').text
    proposer_text = soup.find('a',class_='sc-7yzmni-9 koQguv sc-1vobn0f-10 kvTyxp').text
    
    row = sheet.max_row+1
    sheet.cell(row=row, column=1).value = number
    sheet.cell(row=row, column=2).value = topic
    sheet.cell(row=row, column=3).value = voting_dates  
    sheet.cell(row=row, column=4).value = for_votes
    sheet.cell(row=row, column=5).value = against_votes
    sheet.cell(row=row, column=6).value = details_text
    sheet.cell(row=row, column=7).value = description_text
    sheet.cell(row=row, column=8).value = proposer_text
    
    workbook.save(outfile_file)
    
    row += 1

    browser.quit()
    
    
directory_path = 'data'
outfile_file = os.path.join(directory_path,'data.xlsx')

# Load the data file
workbook = openpyxl.load_workbook(outfile_file)
input_sheet = workbook['Portal Summary Extract']

# Create a new workbook for posts

output_sheet = workbook.create_sheet(title='Portal Complete Extract')
# create titles

output_sheet.cell(row=1, column=1).value = 'Number'
output_sheet.cell(row=1, column=2).value = 'Topic'
output_sheet.cell(row=1, column=3).value = 'Voting end date'
output_sheet.cell(row=1, column=4).value = 'For votes'
output_sheet.cell(row=1, column=5).value = 'Against votes'
output_sheet.cell(row=1, column=6).value = 'Details'
output_sheet.cell(row=1, column=7).value = 'Description'
output_sheet.cell(row=1, column=8).value = 'Proposer'
workbook.save(outfile_file)

# Iterate through the links in the data file
for row in input_sheet.iter_rows(min_row=2, max_col=4, values_only=True):
    number,topic, link = row[0],row[1], row[3]
    print(f'{number} - {topic} started')
    scrape_posts(number,link, topic, output_sheet)
    print(f'{number} - {topic} finished')
    print("*"*50)
    workbook.save(outfile_file)
    
    




