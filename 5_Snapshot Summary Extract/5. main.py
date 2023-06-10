from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl,re,time,os

# Initiate browser driver and open page of interest
geckodriver_path = './webdriver/geckodriver.exe'  # Update with the correct file name and extension
directory_path = 'data'
outfile_file = os.path.join(directory_path,'data.xlsx')

# Check/create output directory
if not os.path.exists(directory_path):
    os.makedirs(directory_path)

browser = webdriver.Firefox(executable_path=geckodriver_path)
browser.get("https://snapshot.org/#/uniswap")

# Infinite scroll till the end of the page
first_run = 1
while True:
    # Execute JavaScript to scroll to the bottom
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    # Wait for some time for the new content to load
    # You can adjust the sleep duration based on the website and your internet speed
    time.sleep(5)  # Wait for 2 seconds

    # Check if the bottom of the page is reached
    current_height = browser.execute_script("return document.body.scrollHeight")
    if first_run ==1:
        prev_height = 0
        first_run =0
    if current_height == prev_height:
        break  # Reached the end of the page, exit the loop
    else:
        prev_height = current_height
        
print("Done scrolling", "\n","*"*15)

# Get the HTML source of the page and parse using bs4
html_source = browser.page_source

soup = BeautifulSoup(html_source, 'html.parser')

# Open Workbook object and create titles
workbook = openpyxl.load_workbook(outfile_file)
sheet = workbook.create_sheet('Snapshot Summary')
sheet.cell(row=1, column=1).value = 'Proposer'
sheet.cell(row=1, column=2).value = 'Title'
sheet.cell(row=1, column=3).value = 'Excerpt Description'
sheet.cell(row=1, column=4).value = 'Status'
sheet.cell(row=1, column=5).value = 'Ending Date'
sheet.cell(row=1, column=6).value = 'Link'
sheet.cell(row=1, column=7).value = 'Option 1'
sheet.cell(row=1, column=8).value = 'Option 2'
sheet.cell(row=1, column=9).value = 'Option 3'

workbook.save(outfile_file)

# Perform data extraction using BeautifulSoup methods
data = soup.find_all('div', class_='border-y border-skin-border bg-skin-block-bg text-base md:rounded-xl md:border transition-colors')
row = 2
for discussion in data:
    proposer = discussion.find('span', class_='w-full cursor-pointer truncate text-skin-link').text
    title = discussion.find('h3', class_='inline pr-2').text
    try:
        excerpt = discussion.find('p', class_='mb-2 line-clamp-2 break-words text-md').text[:50]
    except:
        excerpt = ""
    status = discussion.find('span', class_='bg-violet-600 State text-white').text
    ending_date = discussion.find('span', class_='cursor-help text-sm').text
    link = 'https://snapshot.org/'+discussion.find('a',class_="")['href']
    options = discussion.find_all('div', class_='relative mt-1 w-full')
    count = 0
    for number,option in enumerate(options):
        try:
            option_text = option.find('div', class_='absolute ml-3 flex items-center leading-[43px] text-skin-link').text
            option_name = ' '.join(option_text.split()[:-2])
            option_coin = option.find('span',class_='ml-1 text-skin-text').text
            option_percentage = option.find('div', class_='absolute right-0 mr-3 leading-[40px] text-skin-link').text
            option_details = (option_name,option_coin,option_percentage)
            sheet.cell(row=row, column=7+count).value = f"{option_name}, {option_coin}, {option_percentage}"
            count+=1
        except:
            pass

    sheet.cell(row=row, column=1).value = proposer
    sheet.cell(row=row, column=2).value = title
    sheet.cell(row=row, column=3).value = excerpt
    sheet.cell(row=row, column=4).value = status
    sheet.cell(row=row, column=5).value = ending_date
    sheet.cell(row=row, column=6).value = link
    workbook.save(outfile_file)
    
    row+=1
    
    
    
