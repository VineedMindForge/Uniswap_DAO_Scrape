from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
# from selenium.webdriver.common.action_chains import ActionChains

from bs4 import BeautifulSoup
import openpyxl,re,time,os

# Initiate browser driver and open page of interest
geckodriver_path = './webdriver/geckodriver.exe'  # Update with the correct file name and extension


# # Output file configuration
output_directory_path = 'data'
outfile_file = os.path.join(output_directory_path,'data.xlsx')
# workbook = openpyxl.Workbook()
# sheet = workbook.active
# sheet.title = 'Portal Summary Extract'



def open_or_create_workbook():
    try:
        # Try to open the existing workbook
        workbook = openpyxl.load_workbook(outfile_file)
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        workbook = openpyxl.Workbook()
    
    # Check if the 'Portal Summary Extract' sheet already exists
    if 'Portal Summary Extract' not in workbook.sheetnames:
        # Create a new sheet called 'Portal Summary Extract'
        sheet = workbook.create_sheet('Portal Summary Extract')
    else:
        # Get the existing 'Portal Summary Extract' sheet
        sheet = workbook['Portal Summary Extract']
    
    return workbook,sheet

workbook,sheet = open_or_create_workbook()

sheet.cell(row=1, column=1).value = 'Number'
sheet.cell(row=1, column=2).value = 'Proposal'
sheet.cell(row=1, column=3).value = 'Status'
sheet.cell(row=1, column=4).value = 'Link'

# Check/create output directory
if not os.path.exists(output_directory_path):
    os.makedirs(output_directory_path)

browser = webdriver.Firefox(executable_path=geckodriver_path)
browser.get("https://app.uniswap.org/#/vote")

wait = WebDriverWait(browser, 10)
toggle_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.sc-1poje5t-0')))

browser.execute_script("arguments[0].click();", toggle_button)

# Infinite scroll till the end of the page
print("Loading full page by scrolling")
first_run = 1
while True:
    # Execute JavaScript to scroll to the bottom
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    # Wait for some time for the new content to load
    # You can adjust the sleep duration based on the website and your internet speed
    time.sleep(2)  # Wait for 2 seconds

    # Check if the bottom of the page is reached
    current_height = browser.execute_script("return document.body.scrollHeight")
    if first_run ==1:
        prev_height = 0
        first_run =0
    if current_height == prev_height:
        break  # Reached the end of the page, exit the loop
    else:
        # print(f"Current Height - {current_height}\t Prev Height - {prev_height}")
        prev_height = current_height
        # print('Moving down')

print("Done scrolling", "\n","*"*15)

# Get the HTML source of the page and parse using bs4
html_source = browser.page_source

soup = BeautifulSoup(html_source, 'html.parser')

data = soup.find_all('a', class_='sc-uxt6ak-2 dgkaVU')

row = 2
for proposal in data:
    number = proposal.find('span', class_="sc-uxt6ak-3 gyFpxT").text
    article = proposal.find('span', class_='sc-uxt6ak-4 Dzbgo').text
    status = proposal.find('span',class_=lambda x: x and x.startswith('sc-1z0b5a1')).text
    link = 'https://app.uniswap.org/' + proposal['href']
    
    sheet.cell(row=row, column=1).value = number
    sheet.cell(row=row, column=2).value = article
    sheet.cell(row=row, column=3).value = status
    sheet.cell(row=row, column=4).value = link
    
    row+=1
    
    workbook.save(outfile_file)

browser.quit()  
    

